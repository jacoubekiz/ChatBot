from rest_framework.views import APIView
from django.views.decorators.csrf import csrf_exempt
from rest_framework.generics import GenericAPIView, ListCreateAPIView, RetrieveDestroyAPIView, RetrieveUpdateDestroyAPIView, DestroyAPIView, UpdateAPIView, get_object_or_404
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q
from django.core.files.storage import default_storage
from api.filters import ContactFilter
from django_filters.rest_framework import DjangoFilterBackend
from api.tasks import send_whatsapp_campaign
from .models import *
from .configure_api import *
from .handel_time import *
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import *
from .utils import *
import datetime
from .permissions import UserIsAdmin
from rest_framework.generics import ListAPIView, RetrieveAPIView
# from django_redis import get_redis_connection
from django.utils.decorators import method_decorator
import threading
from .send_email import *
import openpyxl
from .pagination import *
from django.http import HttpResponse
import pandas as pd
import csv
from celery import shared_task

@shared_task
def write_inside_excel(data):
        response = data['response']
        time_meeting = data['time_meeting']
        phonenumber = data['phonenumber']
        try:
            workbook =  openpyxl.load_workbook('media/Seamless_DXB.xlsx')
            
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'client name'
        sheet['B1'] = 'phone number'
        sheet['C1'] = 'time_meeting'
        last_row = sheet.max_row
        new_data = [response, phonenumber, time_meeting]
        sheet.cell(row=last_row + 1, column=1, value=new_data[0])
        sheet.cell(row=last_row + 1, column=2, value=new_data[1])
        sheet.cell(row=last_row + 1, column=3, value=new_data[2])
        workbook.save('media/Seamless_DXB.xlsx')

class RegisterResponseClient(APIView):
    def post(self, request):
        data = request.data
        # thread = threading.Thread(target=write_inside_excel, args=(data,))
        # thread.start()
        write_inside_excel.delay(data)
        return Response(status=status.HTTP_200_OK)

# class ClientsViewSet(viewsets.ModelViewSet):
#     serializer_class = ClientSerializer
#     queryset = Client.objects.all()


# class CreateCalenderView(GenericAPIView):
#     permission_classes = [IsAuthenticated]

#     def post(self, request):
#         data = request.data
#         serializer = CalenderSerializer(data=data,  context={'request':request})
#         serializer.is_valid(raise_exception=True)
#         serializer.save()
#         return Response(serializer.data ,status=status.HTTP_201_CREATED)
    
# class GetCalenderView(GenericAPIView):

#     def get(self, request, user_id):
#         calendar = Calendar.objects.filter(user__id=user_id).all()
#         serializer = CalenderSerializer(calendar, many=True)
#         book_an_appointment = BookAnAppointment.objects.filter(Q(user__id = user_id) & Q(day__day__gte=timezone.now().day))
#         serializer_book = BookAnAppointmentSerializer(book_an_appointment, many=True)
#         return Response({'calender':serializer.data, 'busy_tiem':serializer_book.data}, status=status.HTTP_200_OK)
    
# class CreateWorkingTimeView(ListCreateAPIView):
#     queryset = WorkingTime.objects.all()
#     serializer_class = WorkingTimeSerializer
#     permission_classes = [IsAuthenticated]

#     def get_serializer_context(self):
#         context = super().get_serializer_context()
#         context['request'] = self.request
#         return context

# # class CreateListWorkingHoursPMView(ListCreateAPIView):
# #     queryset = WorkingHoursPM.objects.all()
# #     serializer_class = WorkingHoursPMSerializer
# #     permission_classes = [IsAuthenticated]


# class CreateBookAnAppointmentView(ListCreateAPIView):
#     queryset = BookAnAppointment.objects.all()
#     serializer_class = BookAnAppointmentSerializer

# class GetCalendarForUserView(GenericAPIView):

#     def get(self, request, calender_key):
        
#         calendar = Calendar.objects.filter(key=calender_key).all()
#         calendar_serializer = CalenderSerializer(calendar, many=True)
#         working_days = []
#         calendars = []
#         for calendar_user in calendar_serializer.data:
#             list_working_days = calendar_user['working_time']
#             duration = Duration.objects.get(id=calendar_user['duration'])
#             calendars.append({'calendar_id':calendar_user['id'],"api-key":calendar_user['key'], 'duration':duration_string(duration.duration)})

#         for work_day in list_working_days:
#             working_days.append(work_day['day'])

#         data = {
#             'working_days':working_days,
#             'calendar':calendars
#             }
#         return Response(data , status=status.HTTP_200_OK)
    

# class GetHoursFree(GenericAPIView):

#     def get(self, request):
#         days = ['الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت']
#         # info = request.data
#         free_hours = []
#         day_number = days.index(get_day_name(request.GET.get('date')))
#         if day_number == 0:
#             day_number +=1
#         calendar = Calendar.objects.filter(key=request.GET.get('key')).first()
#         working_time = calendar.working_time.get(day=day_number)
#         start_work_am = convert_time_to_timedelta(working_time.starting_time_am)
#         start_work_pm = convert_time_to_timedelta(working_time.starting_time_pm)
#         end_work_am = convert_time_to_timedelta(working_time.end_time_am)
#         end_work_pm = convert_time_to_timedelta(working_time.end_time_pm)
#         duration = calendar.duration.duration
#         time_slots = []

#         user_book_an_appointment = calendar.user.bookanappointment_set.filter(Q(day=request.GET.get('date'))).order_by('day', 'hour')
#         if not user_book_an_appointment:
#             free_hours.append((convert_timedelta_to_time(start_work_am), convert_timedelta_to_time(end_work_am)))
#             free_hours.append((convert_timedelta_to_time(start_work_pm), convert_timedelta_to_time(end_work_pm)))
#             for i in free_hours:
#                 time_slots.extend(split_time(i[0], i[1], duration))
#             return Response({'free_hours':time_slots})

#         starting_appointment = []
#         end_appointment = []
        
#         for x in user_book_an_appointment:
#             starting_appointment.append((convert_time_to_timedelta(x.hour)))
#             end_appointment.append(convert_time_to_timedelta(x.hour) + x.duration)

#         starting_appointment.insert(0 , start_work_am)
#         starting_appointment.append(end_work_pm)
#         end_appointment.insert(0, start_work_am)
#         end_appointment.append(end_work_pm)
#         # deferance = []
#         for item in range(len(starting_appointment)):
#             try:
#                 next_appointment = starting_appointment[item + 1]
#             except:
#                 next_appointment = starting_appointment[item]
#             if next_appointment-end_appointment[item] >= duration:
#                 free_hours.append((convert_timedelta_to_time(end_appointment[item]), convert_timedelta_to_time(next_appointment)))

#         for free in free_hours:
#             if convert_time_to_timedelta(free[0]) <= end_work_am and convert_time_to_timedelta(free[1]) >= end_work_am:
#                 free_hours.insert(free_hours.index(free), (free[0], convert_timedelta_to_time(end_work_am)))
                
#                 if convert_time_to_timedelta(free[1]) - start_work_pm >= duration:
#                     free_hours.insert(free_hours.index(free)+1, (convert_timedelta_to_time(start_work_pm), free[1]))
#                 free_hours.remove(free)

#         for i in free_hours:
#             time_slots.extend(split_time(i[0], i[1], duration))

#         return Response({'free_hours':time_slots},status=status.HTTP_200_OK)

# class GetFirstTenDays(APIView):

#     def get(self, request):
#         days = ['الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت']
        
#         # info = request.data
#         calendar = Calendar.objects.filter(key=request.GET.get('key')).first()
#         free_hours = []
#         free_days = set()
#         next_days = []
#         if request.GET.get('date') == '':
#             day = calendar.start_appointment
#         else:
#             day = datetime.datetime.strptime(request.GET.get('date'), '%Y-%m-%d').date()
#         # while len(free_days) <= 10
        
#         for d in range(15):
#             if day + timedelta(days=d) <= calendar.end_appointment :
#                 next_days.append(day + timedelta(days=d))
#         for next_day in next_days:
#             if len(free_days) == 9:
#                 break
#             day_number = days.index(get_day_name(next_day))
#             if day_number == 0:
#                 day_number +=1
#             if day_number == 6 or day_number == 5:
#                 continue
            
#             working_time = calendar.working_time.get(day=day_number)
#             start_work_am = convert_time_to_timedelta(working_time.starting_time_am)
#             end_work_am = convert_time_to_timedelta(working_time.end_time_am)
#             end_work_pm = convert_time_to_timedelta(working_time.end_time_pm)
#             duration = calendar.duration.duration
#             user_book_an_appointment = calendar.user.bookanappointment_set.filter(Q(day=next_day)).order_by('day', 'hour')
#             if not user_book_an_appointment:
#                 free_days.add( next_day)
#                 continue
#             starting_appointment = []
#             end_appointment = []
            
#             for x in user_book_an_appointment:
#                 starting_appointment.append((convert_time_to_timedelta(x.hour)))
#                 end_appointment.append(convert_time_to_timedelta(x.hour) + x.duration)

#             starting_appointment.insert(0 , start_work_am)
#             starting_appointment.append(end_work_pm)
#             end_appointment.insert(0, start_work_am)
#             end_appointment.append(end_work_pm)
#             for item in range(len(starting_appointment)):
#                 try:
#                     next_appointment = starting_appointment[item + 1]
#                 except:
#                     next_appointment = starting_appointment[item]
#                 if next_appointment-end_appointment[item] >= duration:
#                     print(next_day)
#                     free_days.add(next_day)
#                     continue

#             for free in free_hours:
#                 if convert_time_to_timedelta(free[0]) <= end_work_am and convert_time_to_timedelta(free[1]) >= end_work_am:
#                     free_days.add(next_day)
#                     continue
   
#         return Response({'free_days':sorted(list(free_days))},status=status.HTTP_200_OK)


# # class GetDoctorsView(APIView):
# #     def get(self, request):
# #         user = CustomUser.objects.all()
# #         serializer_user = SerializerSignUp(user, many=True)
# #         data = serializer_user.data

# #         return Response({'username':data['username']}, status=status.HTTP_200_OK)
    
# class GetDoctorsCalanderView(APIView):
#     def get(self, request, doctor_id):
#         user = CustomUser.objects.get(id=doctor_id)
#         calander = user.calendar_set.all()
#         durations = []
#         for cal in calander:
#             durations.append(convert_timedelta_to_time(cal.duration.duration))

#         return Response({'duration':durations}, status=status.HTTP_200_OK)
    



# class SendEmailView(APIView):
#     def post(self, request):
#         data= {'to_email':request.data['email'], 'email_subject':'','message': request.data['message']}
#         Utlil.send_email(data)
#         return Response(status=status.HTTP_200_OK)
# --------------------------------------------------------------------------------------------------------------
class ListCreateTeamView(ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Team.objects.all()
    serializer_class = TeamSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['account_id'] = self.kwargs['account_id']
        return context
    
    def get_queryset(self):
        account_id = Account.objects.get(account_id=self.kwargs['account_id'])
        return account_id.team_set.all()

class RetrieveUpdateDeleteTeamView(RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TeamSerializer
    lookup_field = 'team_id'
    def get_queryset(self):
        account_id = self.kwargs['account_id']
        team_id = self.kwargs['team_id']
        return Team.objects.filter(account_id=account_id, team_id=team_id)
    
    def perform_update(self, serializer):
        account_id = Account.objects.get(account_id=self.kwargs['account_id'])
        serializer.save(account_id=account_id)
    
class AssigningPermissions(APIView):
    def post(self, request, user_id):
        user = CustomUser.objects.get(id=user_id)
        role = request.data['role']
        add = request.GET.get('add')
        content_type = ContentType.objects.get_for_model(CustomUser)
        permission = Permission.objects.get(
            codename= role,
            content_type=content_type
        )
        if add == 'True':
            user.user_permissions.add(permission)
        else:
            user.user_permissions.remove(permission)
        return Response(status=status.HTTP_200_OK)
    

class ListTeamMember(APIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TeamMemberSerializer
    def get(self, request, team_id):
        team = get_object_or_404(Team, team_id=team_id)
        members = team.members.all()
        serializer = self.serializer_class(members, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class CreateTeamMemberView(GenericAPIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, account_id):
        data_request = request.data
        serializer = AddUserSerializer(
            data = data_request, 
            many=False, 
            context={
                'role':request.data['role'], 
                'account_id': account_id
            }
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
class AddUserForTeam(GenericAPIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request, team_id):
        team = Team.objects.filter(team_id=team_id).first()
        if not team:
            return Response({'error':'Team not found'}, status=status.HTTP_200_OK)
        users = request.data['users']
        keword = request.GET.get('keword')
        for user in users:
            t_user = CustomUser.objects.filter(id=user).first()
            if not t_user:
                return Response({'error':'User not found'}, status=status.HTTP_200_OK)
            if keword == 'add':
                team.members.add(t_user)
            elif keword == 'delete':
                team.members.remove(t_user)
        return Response(status=status.HTTP_200_OK)

class RetrieveUpdateDeleteTeamMemberView(RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = UpdateTeamMemberSerializer
    queryset = CustomUser.objects.all()
    lookup_field = 'pk'

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['role'] = self.request.data.get('role')
        return context
    
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        response_data = serializer.data
        user = get_object_or_404(CustomUser, email=response_data['email'])
        response_data['role'] = [perm.codename for perm in user.user_permissions.all()]
        return Response(response_data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        response_data = serializer.data
        user = get_object_or_404(CustomUser, email=response_data['email'])
        response_data['role'] = [perm.codename for perm in user.user_permissions.all()]
        return Response(response_data)

class CreateListAccount(GenericAPIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        data_request = request.data
        serializer = AddAccountSerializer(data = data_request, many=False)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        email = data_request['email']
        user = get_object_or_404(CustomUser, email=email)
        account = Account.objects.create(
            user=user,
            name=user.username
        )
        return Response(status=status.HTTP_201_CREATED)

    def get(self, request):
        accounts = Account.objects.filter(user__role_user='admin')
        if not accounts:
            return Response({'error':'Dont have permission for this action'}, status=status.HTTP_200_OK)
        serializer = AccontSerializer(accounts, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
class RetrieveUpdateDeleteAccount(RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = UpdateAccountSerializer
    queryset = CustomUser.objects.all()

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['user_id'] = self.kwargs.get('pk')
        return context


class ListCreateChannelView(ListCreateAPIView):
    
    serializer_class = ChannleSerializer
    permission_classes = [IsAuthenticated,]
    def get_queryset(self):
        account_id = self.kwargs['account_id']
        return Channle.objects.filter(account_id=account_id)
    
    def perform_create(self, serializer):
        account_id = Account.objects.get(account_id=self.kwargs['account_id'])
        serializer.save(account_id=account_id)
    
class RetrieveUpdateDeleteChannelView(RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated,]
    serializer_class = ChannleSerializer
    lookup_field = 'account_id'
    def get_queryset(self):
        account_id = self.kwargs['account_id']
        channel = self.kwargs['channel_id']
        return Channle.objects.filter(account_id=account_id, channle_id=channel)

    def perform_update(self, serializer):
        account_id = get_object_or_404(Account, account_id=self.kwargs['account_id'])
        serializer.save(account_id=account_id)

class CreateNewContact(GenericAPIView):
    def post(self, request, account_id, channel_id):
        data = request.data
        account_id = get_object_or_404(Account, account_id=account_id)
        channel_id = get_object_or_404(Channle, channle_id=channel_id)
        contact, created = Contact.objects.get_or_create(phone_number=data['phone_number'], account_id=account_id)
        contact.name = request.data['name']
        contact.save()
        if created:
            conversation = Conversation.objects.create(
                contact_id=contact, 
                channle_id=channel_id, 
                account_id=account_id
                )
            serializer = ContactSerializer(contact, context={'channel_id': channel_id.channle_id, 'account_id': account_id.account_id})
            data = serializer.data
            return Response(data, status=status.HTTP_200_OK)
        else:
            serializer = ContactSerializer(contact, context={'channel_id': channel_id.channle_id, 'account_id': account_id.account_id})
            data = serializer.data
            return Response(data, status=status.HTTP_302_FOUND)
    
class RetrieveUpdateDestroyContactView(RetrieveUpdateDestroyAPIView):
    serializer_class = ContactSerializer
    permission_classes = [IsAuthenticated]
    queryset = Contact.objects.all()
    lookup_field = 'contact_id'

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['account_id'] = self.kwargs['account_id']
        context['channel_id'] = self.kwargs['channel_id']
        return context

class RefreshTokenView(GenericAPIView):
    def post(self, request):
        refresh_token = request.data['refresh']
        token = RefreshToken(refresh_token)
        data = {
            'access': str(token.access_token)
        }
        return Response(data, status=status.HTTP_200_OK)
    
class ViewLogin(GenericAPIView):

    def post(self, request):
        data_request = request.data
        serializer = LoginSerializer(data = data_request, many=False)
        serializer.is_valid(raise_exception=True)
        email = data_request['email']
        try:
            user = get_object_or_404(CustomUser, email=email)
            token = RefreshToken.for_user(user)
            tokens = {'refresh':str(token), 'access':str(token.access_token)}
            team = Team.objects.filter(members__id=user.id).first()
            account_id = team.account_id.account_id
            channel_id = Channle.objects.filter(account_id__account_id = account_id).first()
            if user.role_user == 'admin':
                data = {
                    'tokens':tokens,
                    'user': {
                        'id':user.id,
                        'name':user.username,
                        'role':user.role_user,
                        'account_id': account_id,
                        'channel_id': channel_id.channle_id,
                        'permissions': [perm.split('.')[1] for perm in user.get_all_permissions()]
                    }
                }
            else:
                data = {
                    'tokens':tokens,
                    'user': {
                        'id':user.id,
                        'name':user.username,
                        'permissions':[perm.split('.')[1] for perm in user.get_all_permissions()],
                        'account': {
                            "account_id":account_id,
                            "name": team.account_id.name,
                            "email": team.account_id.user.email,
                            "channel_id":channel_id.channle_id
                        }
                    }
                }
        except:
            user = get_object_or_404(CustomUser, email=email)
            token = RefreshToken.for_user(user)
            tokens = {'refresh':str(token), 'access':str(token.access_token)}
            data = {
                'tokens':tokens,
                'user': {
                    'id':user.id,
                    'name':user.username,
                    'permissions':[perm.split('.')[1] for perm in user.get_all_permissions()],
                }
            }
        return Response(data, status=status.HTTP_200_OK)
    
# End Points for Logout User
class LogoutAPIView(APIView):
    serializer_class = LogoutSerializer
    permission_classes = [IsAuthenticated,]

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({"message": 'logout true'})

class GenerateapiKeyView(GenericAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, account_id):
        account = get_object_or_404(Account, account_id=account_id)
        account.apiKey= account.generate_key()
        account.save()
        message = {
            "account": account.name,
            "apiKey": account.apiKey
        }

        return Response(message, status=status.HTTP_200_OK)

    def get(sefl, request, account_id):
        account = get_object_or_404(Account, account_id=account_id)
        message = {
            "apikey": account.apiKey
        }
        return Response(message, status=status.HTTP_200_OK)

class ListContactView(ListAPIView):
    queryset = Contact.objects.all().prefetch_related(
        'conversation_set__tags'
    )
    serializer_class = ContactSerializerView
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = ContactFilter

class ListConversationView(GenericAPIView):
    serializer_class = ConversationSerializer
    permission_classes = [IsAuthenticated]

    def get(self, request, channel_id):
        user = get_object_or_404(CustomUser, id=request.user.id)
        permissions = list(user.get_all_permissions())
        if 'api.visibility all conversations' in permissions:
            conversation = Conversation.objects.filter(channle_id=channel_id)
            if not conversation:
                return Response({'error':'No conversations found'}, status=status.HTTP_200_OK)
            serializer = self.get_serializer(conversation, many=True, context={'user':request.user})
            return Response(serializer.data)
        else:
            conversation = Conversation.objects.filter(channle_id=channel_id, user=user)
            if not conversation:
                return Response({'error':'No conversations found'}, status=status.HTTP_200_OK)
            serializer = self.get_serializer(conversation, many=True, context={'user':request.user})
            return Response(serializer.data)
    
    def post(self, request, channel_id):
        # data = request.data
        channel = get_object_or_404(Channle, channle_id = channel_id)
        contact = get_object_or_404(Contact, contact_id = channel.contact_id.contact_id)
        conversation, created = Conversation.objects.get_or_create(contact_id = contact , channle_id = channel)
        conversation_serializer = ConverstionSerializerCreate(conversation, many=False)
        return Response(conversation_serializer.data)

class DeleteConversation(DestroyAPIView):
    queryset = Conversation.objects.all()
    serializer_class = ConversationSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'conversation_id'

class ReasignConversation(GenericAPIView):
    def post(self, request, conversation_id):
        data = request.data
        conversation = get_object_or_404(Conversation ,conversation_id=conversation_id)
        user = get_object_or_404(CustomUser, id=data['user_id'])
        conversation.user = user
        conversation.save()

        return Response(status=status.HTTP_200_OK)

class HandelCSView(GenericAPIView):
    def post(self, request):
        data = request.data
        try:
            df = pd.read_csv(data['file'])
            for index, row in df.iterrows():
                if row.get('Status') != 'Open':
                    filename = f'media/error_2.csv'
                    file_exists = os.path.isfile(filename)
                    with open(filename, 'a', encoding='utf-8') as file:
                        writer = csv.writer(file)
                        if not file_exists:
                            writer.writerow(['Phone Number', 'Name', 'Error'])
                        writer.writerow([row.get('Phone Number'), row.get('Name'), 'Invalid Status Open'])
                elif type(row.get('Phone Number')) != int :
                    filename = f'media/error_2.csv'
                    file_exists = os.path.isfile(filename)
                    with open(filename, 'a', encoding='utf-8') as file:
                        writer = csv.writer(file)
                        if not file_exists:
                            writer.writerow(['Phone Number', 'Name', 'Error'])
                        writer.writerow([row.get('Phone Number'), row.get('Name'), 'Invalid Phone Number'])
                elif type(row.get('Phone Dial Code')) != int :
                    filename = f'media/error_2.csv'
                    file_exists = os.path.isfile(filename)
                    with open(filename, 'a', encoding='utf-8') as file:
                        writer = csv.writer(file)
                        if not file_exists:
                            writer.writerow(['Phone Number', 'Name', 'Error'])
                        writer.writerow([row.get('Phone Number'), row.get('Name'), 'Invalid Phone Number']) 

        except:
            return Response({'error':'Invalid file format. Please upload a CSV file.'}, status=status.HTTP_400_BAD_REQUEST)
        return Response(status=status.HTTP_200_OK)

class CreateListCampaignsView(GenericAPIView):
    serializer_class = CampaignsSerializer
    permission_classes = [IsAuthenticated]

    def get(self, request, channel_id):
        channel = get_object_or_404(Channle, channle_id=channel_id)
        campaigns = WhatsAppCampaign.objects.filter(account_id=channel.account_id)
        if not campaigns:
            return Response({'error':'No campaigns found'}, status=status.HTTP_200_OK)
        serializer_campaigns = self.get_serializer(campaigns, many=True)
        data = serializer_campaigns.data

        return Response(data, status=status.HTTP_200_OK)
    
    def post(self, request, channel_id):
        channel = get_object_or_404(Channle, channle_id=channel_id)
        data = request.data
        file = request.data['file']
        content_template = data.get('content_template')
        campaign_name = data.get('campaign_name')
        user = request.user
        whatsappcampaign = WhatsAppCampaign.objects.create(
                    account_id=channel.account_id,
                    name=campaign_name,
                    template_name=data.get('template_name'),
                    # csv_file = request.data['file'],
                    created_by=user,
                )
        df = pd.read_csv(file)
        data_e = {
            'channel':channel.channle_id, 
            'df':df.to_json(orient='records'), 
            'account':channel.account_id.account_id,
            'content_template': content_template,
            'user_id': user.id,
            'language_code': data.get('language_code'),
            'template_name': data.get('template_name'),
            'template_parameters': data.get('template_parameters'),
            'whatsappcampaign': whatsappcampaign.campaign_id
        }
        payload = json.dumps(data_e)
        send_whatsapp_campaign.delay(payload)
        return Response({'campaign_id': whatsappcampaign.campaign_id}, status=status.HTTP_201_CREATED)

class GetCampaignView(GenericAPIView):
    serializer_class = CampaignSerializer_
    permission_classes = [IsAuthenticated]

    def get(self, request, campaign_id):
        campaign = get_object_or_404(WhatsAppCampaign, campaign_id=campaign_id)
        serializer_campaign = self.get_serializer(campaign)
        data = serializer_campaign.data
        return Response(data, status=status.HTTP_200_OK)

class UserProfileView(RetrieveAPIView):
    queryset = CustomUser.objects.all()
    serializer_class = UserProfileSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'

class ListCreateApiView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, account_id):
        account = get_object_or_404(Account, account_id=account_id)
        data = request.data

        # validate data not empty
        if not data:
            return Response(
                {'error':'No data provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        parameters = data.get('parameters', [])
        custome_attrs = data.get('custome_attrs', [])

        # Validate parameters is a list
        if not isinstance(parameters, list):
            return Response(
                {'error':'Parameters must be a list'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        # Validate custome attributes is a list
        if not isinstance(custome_attrs, list):
            return Response(
                {'error':'Custome attributes must be a list'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = APISerializer(
            data=data, 
            context={
                'account':account, 
                'parameters':parameters,
                'custome_attrs': custome_attrs
            }
        )
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'API created successfully'}, status=status.HTTP_201_CREATED)
        return Response({'error':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
    
    def get(self, request, account_id):
        account = get_object_or_404(Account, account_id=account_id)
        api_objects = API.objects.filter(account_id=account).prefetch_related('api_parameter_set')
        if not api_objects.exists():
            return Response(
                {'message': 'No APIs found for this account'}, 
                status=status.HTTP_200_OK
            )
        
        # api_objects = API.objects.filter(account_id=account)

        result = []
        for api_object in api_objects:
            api_parameters = api_object.api_parameter_set.all()
            serializer_api_param = APIParametersSerializer(api_parameters, many=True) 
            serializer = APISerializer(api_object)

            data_dict = serializer.data
            # data_dict = dict(data)
            data_dict['parameters'] = serializer_api_param.data

            result.append(data_dict)
        return Response(result, status=status.HTTP_200_OK)


class GetApiView(GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = APISerializer
    lookup_field = 'api_id'

    def get(self, request, api_id):
        api_object = API.objects.filter(api_id=api_id).first()
        if not api_object:
            return Response({'error':'No API found'}, status=status.HTTP_200_OK)
        api_parameters = Api_parameter.objects.filter(api=api_object)
        serializer_api_param = APIParametersSerializer(api_parameters, many=True)
        serializer = APISerializer(api_object)
        data = serializer.data
        data_dict = dict(data)
        custome_attributes = Custome_attribute.objects.filter(api=api_object)
        seria = SerializerCustomeAttributes(custome_attributes, many=True)
        data_dict['parameters'] = serializer_api_param.data
        data_dict['custome_attrs'] = seria.data

        return Response(data_dict, status=status.HTTP_200_OK)
            
class UpdateApiview(UpdateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = APISerializer
    queryset = API.objects.all()
    lookup_field = 'api_id'

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['parameters'] = self.request.data.get('parameters', [])
        context['custome_attrs'] = self.request.data.get('custome_attrs', [])
        return context
    
class SaveResponse(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, api_id):
        response = request.data['response']
        api = get_object_or_404(API, api_id = api_id)
        api.response = response
        api.save()
        return Response(status=status.HTTP_200_OK)

class DeleteAPIView(DestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = APISerializer
    queryset = API.objects.all()
    lookup_field = 'api_id'

class APILogVeiw(GenericAPIView):
    serializer_class = APILogSerializer
    # queryset = APILog.objects.all()
    permission_classes = [IsAuthenticated]

    def get(self, request, api_id):
        api = API.objects.filter(api_id=api_id).first()
        if not api:
            return Response({'error':'No API found'}, status=status.HTTP_200_OK)
        api_log = APILog.objects.filter(api=api)
        apis_logs = []
        for api_log_ in api_log:
            data = {
                "id":api_log_.apilog_id,
                "status": api_log_.status_request,
                "message":"",
                "created_at":api_log_.created_at,
                "request": {
                    "url": api_log_.api.endpoint,
                    "method":api_log_.api.method,
                    "data":{
                        api_log_.api.body if api_log_.api.body else "",
                    },
                },
                "response": {
                    "status": api_log_.status_request,
                    "message":"",
                    "payload": {
                        str(api_log_.response) if api_log_.response else ""
                    }
                }
            }
            apis_logs.append(data)
        return Response(apis_logs, status=status.HTTP_200_OK)

class DeleteParameterAPIView(DestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ParameterSerializer
    queryset = Parameter.objects.all()
    lookup_field = 'parameter_id'

class ListCreateAttributeView(ListCreateAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, account_id):
        account = Account.objects.filter(account_id=account_id).first()
        if not account:
            return Response({'error':'No account found'}, status=status.HTTP_200_OK)
        data = request.data
        serializer = SerializerAttributes(
            data=data, 
            context={
                'account':account
            }
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(status=status.HTTP_201_CREATED)
    
    def get(self, request, account_id):
        account = Account.objects.filter(account_id=account_id).first()
        if not account:
            return Response({'error':'No account found'}, status=status.HTTP_200_OK)
        attributes = Attribute.objects.filter(account_id=account)
        if not attributes:
            return Response({'error':'No attributes found'}, status=status.HTTP_200_OK)
        serializer = SerializerAttributes(attributes, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
class RetAupDelAttributeView(RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = SerializerAttributes
    lookup_field = 'id'

    def get_queryset(self):
        attribute_id = self.kwargs['id']
        account = self.kwargs['account_id']
        return Attribute.objects.filter(account=account, id=attribute_id)
    
    def perform_update(self, serializer):
        account_id = get_object_or_404(Account, account_id=self.kwargs['account_id'])
        serializer.save(account=account_id)

class CreateListQuickReplyView(GenericAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, account_id):
        data = request.data
        serializer = QuickReplySerializer(data=data, context={'account_id':account_id})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def get(self, request, account_id):
        account = get_object_or_404(Account, account_id=account_id)
        quick_replies = QuickReply.objects.filter(account_id=account)
        serializer = QuickReplySerializer(quick_replies, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class RetrieveUpdateDeleteQuickReplyView(RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = QuickReplySerializer
    lookup_field = 'quickreply_id'

    def get_queryset(self):
        account = self.kwargs['account_id']
        quick_reply_id = self.kwargs['quickreply_id']
        return QuickReply.objects.filter(account_id=account, quickreply_id=quick_reply_id)
    
    def perform_update(self, serializer):
        account_id = get_object_or_404(Account, account_id=self.kwargs['account_id'])
        serializer.save(account_id=account_id)

class ListCreateTriggerView(GenericAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, account_id):
        data = request.data
        serializer = TriggerSerializer(data=data, context={'account_id':account_id})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def get(self, request, account_id):
        account = get_object_or_404(Account, account_id=account_id)
        triggers = Trigger.objects.filter(account_id=account)
        serializer = TriggerSerializer(triggers, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
class RetrieveUpdateDeleteTriggerView(RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TriggerSerializer
    lookup_field = 'id'

    def get_queryset(self):
        trigger_id = self.kwargs['id']
        account = self.kwargs['account_id']
        return Trigger.objects.filter(id=trigger_id, account_id=account)

    def perform_update(self, serializer):
        account = get_object_or_404(Account, account_id=self.kwargs['account_id'])
        serializer.save(account=account)


class ListReportView(ListAPIView):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]


class ListReportView(RetrieveAPIView):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]

class ListMessgesForSpecificConversation(APIView):
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPaginatins

    def get(self, request, conversation_id):
        paginator = CustomPaginatins()
        # paginator.page_size = 20
        conversation = Conversation.objects.get(conversation_id=conversation_id)
        messages = conversation.chatmessage_set.all().order_by('-created_at')
        result_page = paginator.paginate_queryset(messages, request)
        messages_serializer = ChatMessageSerializer(result_page, many=True)
        return paginator.get_paginated_response(messages_serializer.data)

@method_decorator(csrf_exempt, name='dispatch')
class WebhookView(APIView):

    def post(self, request):
        try:
            payload = json.dumps(request.data)
            handel_request_redis.delay(payload)
            return Response(status=status.HTTP_200_OK)
        except Exception as e:
            f = open('redis_error.txt', 'a')
            f.write(f"Error processign webhok: {str(e)}" + '\n')
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    def get(self, request):
        try:
            data = request.data
            handel_request_redis.delay(data)
            return HttpResponse(data, content_type="text/html")
        except Exception as e:
            f = open('redis_error.txt', 'a')
            f.write(f"Error processign webhok: {str(e)}" + '\n')
            return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)

import base64
from django.http import JsonResponse
class ImageToBase64View(APIView):
    def get(self, request):
        image = request.data['image']
        img_data = image.read()
        encoded_img = base64.b64encode(img_data).decode('utf-8')

        return JsonResponse({
            # "success": True,
            "base64_image": encoded_img
        })
    


class ListAllTeamMembers(GenericAPIView):

    permission_classes = [IsAuthenticated]
    def get(self, request, account_id):
        account = get_object_or_404(Account, account_id=account_id)
        member = CustomUser.objects.filter(Q(role_user="agent") & Q(manager=account.user))
        serializer = MemberSerializer(member, many=True)
        members = serializer.data
        return Response(members, status=status.HTTP_200_OK)


class AddListFlows(GenericAPIView):
    
    permission_classes = [IsAuthenticated]
    def post(self, request, channel_id):
        channel = get_object_or_404(Channle, channle_id = channel_id)
        flow = request.data['flow']
        flow_name = request.data['flow_name']
        flow_ = Flow.objects.create(account=channel.account_id, flow=flow, flow_name=flow_name)
        channel.flows.add(flow_)
        channel.save()

        return Response(status=status.HTTP_200_OK)
    
    def get(self, request, channel_id):
        channel = get_object_or_404(Channle, channle_id = channel_id)
        flows = channel.flows.all()
        serializer = SerializerFlows(flows, many=True, context={'request': request})

        return Response(serializer.data, status=status.HTTP_200_OK)
    

class SetDefaultFlow(GenericAPIView):

    permission_classes = [IsAuthenticated]
    def post(self, request, channel_id):
        data = request.data
        try:
            channel = Channle.objects.filter(channle_id = channel_id).first()
            if not channel:
                return Response({'error':'No Channle found'}, status=status.HTTP_200_OK)
            flows = channel.flows.all()
        except:
            return Response({"error":"Channel matching query dose not exist"})
        for flow in flows:
            if str(flow.id) == data['flow_id']:
                
                flow.is_default = request.GET['is_default']
                flow.save()
            else:
                flow.is_default = 'False'
                flow.save()

        return Response(status=status.HTTP_200_OK)


class UpdateFlowView(GenericAPIView):
    def put(self, request, pk):
        data = request.data
        flow = get_object_or_404(Flow, id=pk)
        flow.flow_name = data['flow_name']
        flow.flow = data['flow']
        flow.save()
        serializer = SerializerFlows(flow, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)
    
class RetrieveFlow(RetrieveDestroyAPIView):
    queryset = Flow.objects.all()
    serializer_class = SerializerFlows
    permission_classes = [IsAuthenticated]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["request"] = self.request
        return context


class InitiateLiveChat(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, conversation_id):
        state = request.data['state']
        conversation = get_object_or_404(Conversation, conversation_id=conversation_id)
        conversation.state = state
        conversation.save()
        return Response(status=status.HTTP_200_OK)


class ChangeConversationStatus(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, conversation_id):
        status_ = request.data['status']
        conversation = get_object_or_404(Conversation, conversation_id=conversation_id)
        conversation.status = status_
        conversation.save()
        return Response(status=status.HTTP_200_OK)

class AddTagToConversation(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, conversation_id):
        tag_ids = request.data.get('tag_ids', [])
        conversation = get_object_or_404(Conversation, conversation_id=conversation_id)
        for tag_id in tag_ids:
            tag = Tag.objects.filter(tag_id=tag_id).first()
            if not tag:
                return Response({"error":f"tag with {tag_id} not found"}, status=status.HTTP_204_NO_CONTENT)
            conversation.tags.add(tag)
        conversation.save()
        return Response(status=status.HTTP_200_OK)

class CreateTagView(GenericAPIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, account_id):
        name = request.data['name']
        account = get_object_or_404(Account, account_id=account_id)
        tag = Tag.objects.create(
            name=name,
            account_id=account
        )
        return Response({'tag_id': tag.tag_id, 'name': tag.name}, status=status.HTTP_201_CREATED)
    
    def get(self, request, account_id):
        account = get_object_or_404(Account, account_id=account_id)
        tags = account.tag_set.all()
        data = []
        for tag in tags:
            data.append({'tag_id': tag.tag_id, 'name': tag.name})
        return Response(data, status=status.HTTP_200_OK)

class ChangePasswordView(GenericAPIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, user_id):
        data = request.data
        user = get_object_or_404(CustomUser,id=user_id)
        serializer = ChangePasswordSerializer(data=data, context={'user': user, 'user_login': request.user})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response({'message': 'Password changed successfully.'}, status=status.HTTP_200_OK)
    

class ListCreateGroupView(ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = GroupSerializer
    queryset = Group.objects.all()
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['account_id'] = self.kwargs['account_id']
        tag = self.request.query_params.get('tag')
        members = Conversation.objects.filter(tags__tag_id=tag).values_list('contact_id', flat=True).distinct()
        if not members:
            return Response({'error':'No members found'}, status=status.HTTP_200_OK)
        context["members"] = members
        return context
    
class RetrieveUpdateDeleteGroupView(RetrieveUpdateDestroyAPIView):
    serializer_class = GroupSerializer
    permission_classes = [IsAuthenticated]
    queryset = Group.objects.all()
    lookup_field = 'id'

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['account_id'] = self.kwargs['account_id']
        tag = self.request.query_params.get('tag')
        members = Conversation.objects.filter(tags__tag_id=tag).values_list('contact_id', flat=True).distinct()
        if not members:
            return {'error':'No members found'}
        context["members"] = members
        return context
    